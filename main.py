#!/usr/bin/env python
# coding: utf-8

# In[1]:


from PyQt5 import QtCore, QtGui, QtWidgets
import sys
from PyQt5 import QtWidgets, uic
from PyQt5.QtGui import QPixmap, QColor

from PyQt5.QtWidgets import QFileDialog, QAction, QTableWidgetItem, QHeaderView

import os
import numpy as np
import pandas as pd

import cv2
import keras
from openpyxl import Workbook

cv2.ocl.setUseOpenCL(True)


# In[2]:


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(569, 568)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(-10, 0, 661, 661))
        self.label.setStyleSheet("background-color: rgb(0, 0, 0);")
        self.label.setText("")
        self.label.setObjectName("label")
        self.ipAdresse = QtWidgets.QLineEdit(self.centralwidget)
        self.ipAdresse.setGeometry(QtCore.QRect(60, 80, 211, 41))
        self.ipAdresse.setAlignment(QtCore.Qt.AlignCenter)
        self.ipAdresse.setObjectName("ipAdresse")
        self.port = QtWidgets.QLineEdit(self.centralwidget)
        self.port.setGeometry(QtCore.QRect(280, 80, 91, 41))
        self.port.setAlignment(QtCore.Qt.AlignCenter)
        self.port.setCursorMoveStyle(QtCore.Qt.VisualMoveStyle)
        self.port.setObjectName("port")
        self.connectButton = QtWidgets.QPushButton(self.centralwidget)
        self.connectButton.setGeometry(QtCore.QRect(380, 80, 111, 41))
        self.connectButton.setObjectName("connectButton")
        self.tableau = QtWidgets.QTableWidget(self.centralwidget)
        self.tableau.setGeometry(QtCore.QRect(20, 200, 531, 301))
        self.tableau.setStyleSheet("background-color: rgb(0, 0, 0);\n"
"font: 11pt \"MV Boli\";")
        self.tableau.setObjectName("tableau")
        self.tableau.setColumnCount(6)
        self.tableau.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableau.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableau.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableau.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableau.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableau.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableau.setHorizontalHeaderItem(5, item)
        self.titreTableau = QtWidgets.QLabel(self.centralwidget)
        self.titreTableau.setGeometry(QtCore.QRect(130, 150, 231, 41))
        self.titreTableau.setStyleSheet("color: rgb(255, 255, 255);\n"
"font: 14pt \"Ravie\";")
        self.titreTableau.setAlignment(QtCore.Qt.AlignCenter)
        self.titreTableau.setObjectName("titreTableau")
        self.exportButton = QtWidgets.QPushButton(self.centralwidget)
        self.exportButton.setGeometry(QtCore.QRect(220, 510, 151, 41))
        self.exportButton.setObjectName("exportButton")
        self.resultConnection = QtWidgets.QLabel(self.centralwidget)
        self.resultConnection.setGeometry(QtCore.QRect(20, 20, 531, 41))
        self.resultConnection.setStyleSheet("color: rgb(255, 0, 0);\n"
"font: 20pt \"Matura MT Script Capitals\";")
        self.resultConnection.setAlignment(QtCore.Qt.AlignCenter)
        self.resultConnection.setObjectName("resultConnection")
        self.listeClasse = QtWidgets.QComboBox(self.centralwidget)
        self.listeClasse.setGeometry(QtCore.QRect(360, 150, 71, 41))
        self.listeClasse.setObjectName("listeClasse")
        self.listeClasse.addItem("")
        self.listeClasse.addItem("")
        self.listeClasse.addItem("")
        self.listeClasse.addItem("")
        self.listeClasse.addItem("")
        self.listeClasse.addItem("")
        self.listeClasse.addItem("")
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.ipAdresse.setPlaceholderText(_translate("MainWindow", "Adresse IP"))
        self.port.setPlaceholderText(_translate("MainWindow", "Port"))
        self.connectButton.setText(_translate("MainWindow", "Connect"))
        item = self.tableau.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Numéro"))
        item = self.tableau.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Noms"))
        item = self.tableau.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Prénoms"))
        item = self.tableau.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "Classe"))
        item = self.tableau.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Note"))
        item = self.tableau.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "Remarque(s)"))
        self.titreTableau.setText(_translate("MainWindow", "Tableau de note"))
        self.exportButton.setText(_translate("MainWindow", "Exporter"))
        self.resultConnection.setText(_translate("MainWindow", "Veuillez connecter la caméra !"))
        self.listeClasse.setItemText(0, _translate("MainWindow", "..."))
        self.listeClasse.setItemText(1, _translate("MainWindow", "LP1"))
        self.listeClasse.setItemText(2, _translate("MainWindow", "L1"))
        self.listeClasse.setItemText(3, _translate("MainWindow", "LP2"))
        self.listeClasse.setItemText(4, _translate("MainWindow", "L2"))
        self.listeClasse.setItemText(5, _translate("MainWindow", "M1"))
        self.listeClasse.setItemText(6, _translate("MainWindow", "MP2"))


# In[3]:


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):

    def resize_specific_column(self, column_index):
        header = self.tableau.horizontalHeader()
        header.setSectionResizeMode(column_index, QHeaderView.ResizeToContents)

    def load_excel_data(self, filename):
        try:
            # Lire le fichier Excel
            df = pd.read_excel(f'Classe/{filename}.xlsx', dtype=str)
            
            # Définir le nombre de lignes et de colonnes dans le QTableWidget
            num_rows, num_cols = df.shape
            self.tableau.setRowCount(num_rows)
            self.tableau.setColumnCount(num_cols)

            # Remplir le QTableWidget avec les valeurs Excel
            for i in range(num_rows - 1):
                for j in range(num_cols):
                    item = QTableWidgetItem(str(df.iloc[i, j]))
                    item.setForeground(QColor("white"))
                    self.tableau.setItem(i, j, item)

            # Nommer les colonnes
            self.tableau.setHorizontalHeaderLabels(df.columns)
            self.tableau.verticalHeader().setVisible(False)

            # Activer le retour à la ligne automatique dans les cellules
            self.resize_specific_column(0)
            self.resize_specific_column(1)

            self.number = num_rows

        except Exception as e:
            print("Une erreur s'est produite lors du chargement des données Excel:", e)

    def __init__(self, *args, obj=None, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.setupUi(self)
        self.initComposant()
        os.system('cls')
        self.model = keras.saving.load_model('final.keras')
        self.number = 1
        self.note = []

    def predire(self, image, model):
        data = []
        image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        resized_image = cv2.resize(image, (100, 100))
        resized_image = cv2.GaussianBlur(resized_image, (5, 5), 0)
        resized_image = cv2.normalize(resized_image, None, 0, 255, cv2.NORM_MINMAX)
        data.append(np.array(resized_image))

        pred = model.predict(np.array(data))
        pred = np.argmax(pred, axis=1)

        return pred[0]

    def initComposant(self):
        self.listeClasse.currentIndexChanged.connect(self.selection_changed)
        self.connectButton.clicked.connect(self.connecter)
        self.exportButton.clicked.connect(self.exporter)

    def selection_changed(self, index):
        selected_option = self.listeClasse.currentText()
        if selected_option == '...':
            self.tableau.setRowCount(0)
        else:
            self.load_excel_data(selected_option)

    def draw_rectangles_and_annotations(self, frame, rectangles):
        for i, (x, y, w, h) in enumerate(sorted(rectangles, key=lambda rect: rect[0]), start=1):
            # Dessiner le rectangle décalé sur l'image
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

            roi = frame[y:y + h, x:x + w]
            titre = self.predire(roi, self.model)

            # Ajouter l'annotation
            cv2.putText(frame, f"{titre}", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 5, (0, 255, 0), 2)
            self.note.append(titre)

    def exporter(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Exporter Excel", "", "Excel Files (*.xlsx)")

        if file_path:
            wb = Workbook()
            ws = wb.active

            # Écrire les noms de colonnes
            for column in range(self.tableau.columnCount()):
                header = self.tableau.horizontalHeaderItem(column)
                ws.cell(row=1, column=column + 1).value = header.text() if header is not None else f"Colonne {column + 1}"

            # Écrire les données de la table
            for row in range(self.tableau.rowCount()):
                for column in range(self.tableau.columnCount()):
                    item = self.tableau.item(row, column)
                    if item is not None:
                        ws.cell(row=row + 2, column=column + 1).value = item.text()

            wb.save(file_path)

    def connecter(self):
        self.resultConnection.setText('Camera connecte')
        phone_ip = self.ipAdresse.text()
        phone_port = self.port.text()

        url = f"http://{phone_ip}:{phone_port}/video"

        cv2.namedWindow('Video', cv2.WINDOW_NORMAL)
        cv2.resizeWindow('Video', 640, 480)  # Taille spécifique (640x480)

        cap = cv2.VideoCapture(url)

        if not cap.isOpened():
            print("Erreur: Impossible de se connecter à la caméra du téléphone.")
            exit()

        green_rectangles = []

        while True:
            # Capturer une image de la caméra du téléphone
            ret, frame = cap.read()

            # Attendre 1 milliseconde pour une touche pressée
            key = cv2.waitKey(1)

            try:
                # Récupérer les dimensions de l'image
                hauteur, largeur, _ = frame.shape

                gray_image = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

                # Appliquer un seuillage pour obtenir une image binaire
                _, threshold_image = cv2.threshold(gray_image, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

                # Trouver les contours dans l'image binaire
                contours, _ = cv2.findContours(threshold_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

                # Vérifier si la touche 'p' est enfoncée pour placer les objets

                # Réinitialiser la liste des rectangles verts

                if key == ord('p'):
                    green_rectangles = []
                    # Filtrer les contours pour récupérer ceux qui correspondent aux chiffres
                    for contour in contours:
                        x, y, w, h = cv2.boundingRect(contour)
                        # Filtrer les contours en fonction de leur taille, forme, etc.
                        if w > 50 and h > 50:
                            # Ajouter les coordonnées du rectangle englobant vert à la liste
                            #if (x > 15):
                             #   x -= 15
                              #  w += 30
                               # h += 30
                            #if (y > 15):
                             #   y -= 15
                              #  w += 30
                               # h += 30
                            green_rectangles.append((x, y, w, h))

                # Vérifier si la touche 'l' est enfoncée pour effacer les objets
                if key == ord('l'):
                    if(self.listeClasse.currentText() == '...'):
                        pass
                    else:
                        noteT = np.array(self.note)
                        print(noteT)
                        
                        nt = 0
                        if (int(f'{noteT[0]}{noteT[1]}') > 20):
                            nt = f'{noteT[0]}{noteT[1]}'
                        else:
                            nt = f'{noteT[0]}{noteT[1]}'

                        for i in range(self.number):
                            if (self.tableau.item(i, 4).text() == '?'):
                                # Réinitialiser la liste des rectangles verts
                                item = QTableWidgetItem(nt)
                                item.setForeground(QColor("white"))
                                self.tableau.setItem(i, 4, item)

                                val = ''
                                data = int(nt)
                                if (data < 10):
                                    val = 'Mediocre'
                                elif (10 <= data < 12):
                                    val = 'Passable'
                                elif (12 <= data < 14):
                                    val = 'Assez-Bien'
                                elif (14 <= data < 16):
                                    val = 'Bien'
                                elif (16 <= data < 18):
                                    val = 'Tres-Bien'
                                else:
                                    val = 'Excellent'

                                item = QTableWidgetItem(val)
                                item.setForeground(QColor("white"))
                                self.tableau.setItem(i, 5, item)
                                break

                    green_rectangles = []
                    self.note = []

                # Dessiner les rectangles englobants verts sur l'image originale
                self.draw_rectangles_and_annotations(frame, green_rectangles)

                # Afficher la vidéo en direct avec les rectangles et les numéros d'objet
                cv2.imshow('Video', frame)

                # Quitter la boucle si la touche 'q' est enfoncée
                if key == ord('q'):
                    self.resultConnection.setText('Camera deconnecte')
                    break
            except Exception as e:
                print(e)
                break

        # Libérer les ressources
        cap.release()
        cv2.destroyAllWindows()


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)

    window = MainWindow()
    window.show()
    app.exec()


# In[ ]:




